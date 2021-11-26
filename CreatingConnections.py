from openpyxl import load_workbook
from shapely.geometry import Point
import pyodbc
import shapely.wkt

# импорт узлов доступа (ОТ)
wb = load_workbook("NODE.xlsx")
ws = wb.active

try:
    # подключение базы полигона и номеров районов
    con_string = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:\Job\pythondb.accdb;"
    conn = pyodbc.connect(con_string)
    print("База полигона и номеров подключена")

    # подключение базы узлов
    con_stroka = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:\Job\Cons.accdb;"
    conn1 = pyodbc.connect(con_stroka)
    print("База узлов подключена", "\n")

    # установка курсоров
    cursor = conn1.cursor()
    # курсор для записи итогов в таблицу
    cur = conn.cursor()
    # курсор для чтения

    cur.execute("SELECT * FROM ZONE")

    records = cur.fetchmany(ws.max_row)

    for row in records:
        P = shapely.wkt.loads(row[4])
        # создание полигона из ячейки E[i]

        # прогон узлов на пересечение, нахождение на границах или внутри площади

        for x in range(1, ws.max_row):
            Point_X = float(ws.cell(row=x, column=2).value)
            # объявление X
            Point_Y = float(ws.cell(row=x, column=3).value)
            # объявление Y
            NODE_ID = int(ws.cell(row=x, column=1).value)
            # объявление узлов
            p1 = Point(Point_X, Point_Y)
            # объявление полигона

            # цикл для проверки попадания точки в район
            if (p1.intersects(P) == True) or (p1.within(P) == True) or (P.contains(p1) == True):
                # print("ZONE_ID",row[0],"NODE_ID",NODE_ID)
                # cursor.execute("SELECT *FROM CONNECTOR") # выборка
                record = cur.fetchmany(1)

                # запись в таблицу примыкания
                cursor.execute("""INSERT INTO CONNECTOR(ZONENO, NODENO, DIRECTION, TSYSSET)
                                    Values(?,?,'D','PR');""", row[0], NODE_ID)
                cursor.execute("""INSERT INTO CONNECTOR(ZONENO, NODENO, DIRECTION, TSYSSET)
                                    Values(?,?,'O','PR')""", row[0], NODE_ID)
                print("Примыкание создано!")
    conn1.commit()
    print("Успешно")
    # print("[", row[0], "]", "[", row[1], "]", "-",)

except pyodbc.Error as e:
    print("Ошибка при подключении баз", e)
