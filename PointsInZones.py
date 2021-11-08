from openpyxl import load_workbook
from shapely.geometry import Point
import pyodbc
import shapely.wkt

# импорт узлов
wb = load_workbook("NODE.xlsx")
ws = wb.active

# импорт полигонов
try:
    con_string = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:\Job\pythondb.accdb;"
    conn = pyodbc.connect(con_string)
    print("Connected", "\n")

    cur = conn.cursor()
    cur.execute("SELECT Name,WKTSURFACE FROM ZONE")

    records = cur.fetchmany(ws.max_row)
    for row in records:
        counter = 0
        P = shapely.wkt.loads(row[1])
        # прогон узлов на пересечение, нахождение на границах или внутри площади
        for x in range(1, ws.max_row):
            Point_X = float(ws.cell(row=x, column=2).value)
            Point_Y = float(ws.cell(row=x, column=3).value)
            p1 = Point(Point_X, Point_Y)

            if (p1.intersects(P) == True) or (p1.within(P) == True) or (P.contains(p1) == True):
                counter += 1
                # print("Yes")
                # print(p1)
        print("Количество узлов доступа в районе ", "[", row[0], "]", "-", counter)

except pyodbc.Error as e:
    print("Error in Connecting", e)
