from openpyxl import load_workbook
from shapely.geometry import Point
import pyodbc
import shapely.wkt

# импорт узлов доступа (ОТ)
wb = load_workbook("NODE.xlsx")
ws = wb.active

try:
    # подключение базы полигона и номеров районов
    con_string = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:\Job\pythondb.accdb;"
    conn = pyodbc.connect(con_string)
    print("База полигона и номеров подключена")

    # подключение базы узлов
    con_stroka = r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:\Job\Cons.accdb;"
    conn1 = pyodbc.connect(con_stroka)
    print("База узлов подключена", "\n")

    # установка курсоров
    cursor = conn1.cursor()
    # курсор для записи итогов в таблицу
    cur = conn.cursor()
    # курсор для чтения

    cur.execute("SELECT * FROM ZONE")

    records = cur.fetchmany(ws.max_row)

    for row in records:
        P = shapely.wkt.loads(row[4])
        # создание полигона из ячейки E[i]
        array_to4ek = [1]
        # прогон узлов на пересечение, нахождение на границах или внутри площади
        print("--- Следующий район ", "[", row[0], "]", "--- ")
        for x in range(1, ws.max_row):
            Point_X = float(ws.cell(row=x, column=2).value)     # объявление X
            Point_Y = float(ws.cell(row=x, column=3).value)     # объявление Y
            NODE_ID = int(ws.cell(row=x, column=1).value)       # объявление узлов
            p1 = Point(Point_X, Point_Y)                        # объявление полигона
            p2 = Point(row[2], row[3])

            # цикл для проверки попадания точки в район
            if (p1.intersects(P) == True) or (p1.within(P) == True) or (P.contains(p1) == True):
                #print("Центр района - ", "[", row[0], "]", "[", row[1], "]","X- ", row[2], "Y- ", row[3])
                #print("В районе ", row[0], "расстояние от центра до узла - ", p1.distance(p2))
                Pmax = p1.distance(p2)
                array_to4ek.append(Pmax)

        if len(array_to4ek) > 1:
            print("Массив точек - ", array_to4ek)
            WorkNode = min(array_to4ek)
            print("Ближайшая к центру точка - ", WorkNode)
        else:
            print("В массиве нет узлов")
        print("--- Переход к следующему району --- ")
        print("\n")

except pyodbc.Error as e:
    print("Ошибка при подключении баз", e)
